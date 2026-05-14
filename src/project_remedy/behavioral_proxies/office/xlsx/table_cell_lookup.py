"""XLSX table-cell lookup proxy over Excel table structures."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from project_remedy.behavioral_proxies.office._checks import report_for, result_from_rules
from project_remedy.behavioral_proxies.shared.base import BehavioralTestResult
from project_remedy.behavioral_proxies.shared.llm_answering import (
    BehavioralAnswerer,
    score_answer_retention,
)
from project_remedy.behavioral_proxies.shared.question_generator import GeneratedQuestion
from project_remedy.models import FileType


@dataclass(frozen=True)
class XLSXTableStructureSignal:
    sheet_name: str
    table_name: str
    ref: str
    row_count: int
    column_count: int
    empty_header_columns: tuple[int, ...]
    has_excel_table: bool
    has_header_row: bool
    has_banded_rows: bool
    has_total_row: bool
    issue: str
    rows: tuple[tuple[str, ...], ...] = ()

    @property
    def passed(self) -> bool:
        return self.issue in {"", "xlsx_table_missing_banded_rows", "xlsx_table_missing_total_row"}

    @property
    def score(self) -> float:
        if self.issue == "":
            return 1.0
        if self.issue in {"xlsx_table_missing_banded_rows", "xlsx_table_missing_total_row"}:
            return 0.85
        return 0.0


class XLSXTableCellLookupTest:
    test_name = "table_cell_lookup"
    dimension = "table_structure"
    format = "xlsx"

    def run(self, artifact_path: Path, **kwargs: Any) -> BehavioralTestResult:
        if artifact_path.exists():
            return _result_from_table_signals(
                xlsx_table_structure_signals(artifact_path),
                test_name=self.test_name,
                dimension=self.dimension,
                fmt=self.format,
                answerer=kwargs.get("answerer"),
                baseline_text=str(kwargs.get("baseline_text") or ""),
                candidate_text=str(kwargs.get("candidate_text") or ""),
            )
        return result_from_rules(
            report_for(artifact_path, FileType.XLSX, kwargs),
            test_name=self.test_name,
            dimension=self.dimension,
            fmt=self.format,
            rule_ids=("xlsx-header-behaviors",),
            threshold=0.95,
        )


def xlsx_table_structure_signals(artifact_path: Path) -> list[XLSXTableStructureSignal]:
    """Return deterministic structure signals for workbook data regions and tables."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        return []

    try:
        workbook = load_workbook(str(artifact_path), data_only=True)
    except Exception:  # noqa: BLE001 - malformed or unsupported workbook.
        return []

    try:
        signals: list[XLSXTableStructureSignal] = []
        for worksheet in workbook.worksheets:
            tables = list(getattr(worksheet.tables, "values", lambda: [])())
            has_data_range = _worksheet_has_data_range(worksheet)
            if has_data_range and not tables:
                signals.append(
                    XLSXTableStructureSignal(
                        sheet_name=worksheet.title,
                        table_name="",
                        ref=_worksheet_used_ref(worksheet),
                        row_count=worksheet.max_row,
                        column_count=worksheet.max_column,
                        empty_header_columns=(),
                        has_excel_table=False,
                        has_header_row=False,
                        has_banded_rows=False,
                        has_total_row=False,
                        issue="xlsx_data_range_missing_excel_table",
                        rows=_worksheet_rows(worksheet, 1, 1, worksheet.max_column, worksheet.max_row),
                    )
                )
                continue
            for table in tables:
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                row_count = max_row - min_row + 1
                column_count = max_col - min_col + 1
                has_header_row = bool(getattr(table, "headerRowCount", 1))
                empty_header_columns: tuple[int, ...] = ()
                if has_header_row:
                    empty_header_columns = tuple(
                        offset + 1
                        for offset, column in enumerate(range(min_col, max_col + 1))
                        if _column_has_data(worksheet, column, min_row + 1, max_row)
                        and not _cell_text(worksheet.cell(row=min_row, column=column))
                    )
                style = getattr(table, "tableStyleInfo", None)
                has_banded_rows = bool(getattr(style, "showRowStripes", False))
                has_total_row = bool(getattr(table, "totalsRowCount", None) or getattr(table, "totalsRowShown", None))
                signals.append(
                    XLSXTableStructureSignal(
                        sheet_name=worksheet.title,
                        table_name=str(getattr(table, "displayName", "") or getattr(table, "name", "")),
                        ref=str(table.ref),
                        row_count=row_count,
                        column_count=column_count,
                        empty_header_columns=empty_header_columns,
                        has_excel_table=True,
                        has_header_row=has_header_row,
                        has_banded_rows=has_banded_rows,
                        has_total_row=has_total_row,
                        issue=_table_issue(
                            row_count=row_count,
                            column_count=column_count,
                            has_header_row=has_header_row,
                            empty_header_columns=empty_header_columns,
                            has_banded_rows=has_banded_rows,
                            has_total_row=has_total_row,
                        ),
                        rows=_worksheet_rows(worksheet, min_col, min_row, max_col, max_row),
                    )
                )
        return signals
    finally:
        workbook.close()


def _result_from_table_signals(
    signals: list[XLSXTableStructureSignal],
    *,
    test_name: str,
    dimension: str,
    fmt: str,
    answerer: BehavioralAnswerer | None = None,
    baseline_text: str = "",
    candidate_text: str = "",
) -> BehavioralTestResult:
    if not signals:
        return BehavioralTestResult(
            test_name=test_name,
            dimension=dimension,
            format=fmt,
            passed=True,
            score=1.0,
            threshold=0.95,
            confidence=0.45,
            metadata={
                "applicable": False,
                "parser_support": "openpyxl_tables",
                "table_count": 0,
                "data_region_count": 0,
                "llm_answering_enabled": answerer is not None,
            },
        )

    failures = [signal for signal in signals if not signal.passed]
    score = sum(signal.score for signal in signals) / len(signals)
    structural_score = score
    lookup_questions = [
        question
        for signal in signals
        for question in _lookup_questions(signal)
    ][:5]
    candidate_context = candidate_text or "\n\n".join(
        _serialize_table(signal) for signal in signals
    )
    findings = [_finding(signal) for signal in signals if signal.issue]
    metadata = {
        "applicable": True,
        "parser_support": "openpyxl_tables",
        "table_count": sum(1 for signal in signals if signal.has_excel_table),
        "data_region_count": len(signals),
        "failure_count": len(failures),
        "llm_answering_enabled": answerer is not None,
        "lookup_question_count": len(lookup_questions),
    }
    if answerer is not None and lookup_questions:
        retention = score_answer_retention(
            questions=lookup_questions,
            baseline_context=baseline_text or candidate_context,
            candidate_context=candidate_context,
            answerer=answerer,
        )
        score = min(structural_score, retention.retention)
        findings.extend(retention.findings)
        metadata.update(
            {
                "baseline_accuracy": retention.baseline_accuracy,
                "candidate_accuracy": retention.candidate_accuracy,
                "answer_accuracy_retention": retention.retention,
            }
        )
    return BehavioralTestResult(
        test_name=test_name,
        dimension=dimension,
        format=fmt,
        passed=score >= 0.95,
        score=round(score, 4),
        threshold=0.95,
        confidence=0.70,
        findings=findings,
        metadata=metadata,
    )


def _table_issue(
    *,
    row_count: int,
    column_count: int,
    has_header_row: bool,
    empty_header_columns: tuple[int, ...],
    has_banded_rows: bool,
    has_total_row: bool,
) -> str:
    if row_count < 2 or column_count < 1:
        return "xlsx_table_too_small_for_lookup"
    if not has_header_row:
        return "xlsx_table_missing_header_row"
    if empty_header_columns:
        return "xlsx_table_empty_header_cells"
    if not has_banded_rows:
        return "xlsx_table_missing_banded_rows"
    if not has_total_row:
        return "xlsx_table_missing_total_row"
    return ""


def _finding(signal: XLSXTableStructureSignal) -> dict[str, Any]:
    severity = "warning" if signal.passed else "error"
    return {
        "severity": severity,
        "issue": signal.issue,
        "sheet_name": signal.sheet_name,
        "table_name": signal.table_name,
        "ref": signal.ref,
        "row_count": signal.row_count,
        "column_count": signal.column_count,
        "empty_header_columns": list(signal.empty_header_columns),
    }


def _worksheet_has_data_range(worksheet: Any) -> bool:
    if worksheet.max_row < 2 or worksheet.max_column < 2:
        return False
    non_empty = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if _cell_text(cell):
                non_empty += 1
                if non_empty >= 2:
                    return True
    return False


def _worksheet_used_ref(worksheet: Any) -> str:
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return ""
    return f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"


def _worksheet_rows(
    worksheet: Any,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
) -> tuple[tuple[str, ...], ...]:
    return tuple(
        tuple(
            _cell_text(worksheet.cell(row=row_index, column=column_index))
            for column_index in range(min_col, max_col + 1)
        )
        for row_index in range(min_row, max_row + 1)
    )


def _column_has_data(worksheet: Any, column: int, min_row: int, max_row: int) -> bool:
    for row in range(min_row, max_row + 1):
        if _cell_text(worksheet.cell(row=row, column=column)):
            return True
    return False


def _cell_text(cell: Any) -> str:
    value = getattr(cell, "value", "")
    return "" if value is None else str(value).strip()


def _lookup_questions(signal: XLSXTableStructureSignal) -> list[GeneratedQuestion]:
    if len(signal.rows) < 2:
        return []
    headers = [cell for cell in signal.rows[0] if cell]
    if not headers:
        return []
    questions: list[GeneratedQuestion] = []
    label = signal.table_name or signal.sheet_name
    for row in signal.rows[1:]:
        if len(row) < 2:
            continue
        row_label = row[0]
        if not row_label:
            continue
        for column_index, expected in enumerate(row[1:], start=1):
            if not expected:
                continue
            header = headers[column_index] if column_index < len(headers) else f"column {column_index + 1}"
            questions.append(
                GeneratedQuestion(
                    question=(
                        f"In XLSX table {label}, what is the value for row "
                        f"{row_label}, column {header}?"
                    ),
                    expected_answer=expected,
                    source_dimension="table_structure",
                )
            )
    return questions


def _serialize_table(signal: XLSXTableStructureSignal) -> str:
    lines = []
    for row in signal.rows:
        values = [cell for cell in row if cell]
        if values:
            lines.append(" | ".join(values))
    return "\n".join(lines)
