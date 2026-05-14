"""XLSX sheet navigation proxy over workbook tabs and sheet contents."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from project_remedy.behavioral_proxies.shared.base import BehavioralTestResult
from project_remedy.behavioral_proxies.shared.llm_answering import (
    BehavioralAnswerer,
    score_answer_retention,
)
from project_remedy.behavioral_proxies.shared.question_generator import GeneratedQuestion


_STOPWORDS = {
    "and",
    "data",
    "details",
    "info",
    "information",
    "sheet",
    "table",
    "the",
    "workbook",
}
_OVERVIEW_SHEET_TOKENS = {
    "contents",
    "dashboard",
    "index",
    "overview",
    "summary",
}
_ISSUE_SCORES = {
    "non_descriptive_sheet_name": 0.0,
    "sheet_name_purpose_unclear": 0.5,
    "overview_sheet_not_first": 0.5,
    "data_sheet_hidden": 0.75,
}


@dataclass(frozen=True)
class XLSXSheetNavigationSignal:
    sheet_name: str
    sheet_index: int
    state: str
    non_empty_cell_count: int
    content_terms: tuple[str, ...]
    issue: str
    issues: tuple[str, ...] = ()

    @property
    def score(self) -> float:
        issues = self.issues or ((self.issue,) if self.issue else ())
        if not issues:
            return 1.0
        return min(_ISSUE_SCORES.get(issue, 0.0) for issue in issues)


class XLSXSheetNavigationTest:
    test_name = "sheet_navigation"
    dimension = "sheet_organization"
    format = "xlsx"

    def run(self, artifact_path: Path, **kwargs: Any) -> BehavioralTestResult:
        signals = xlsx_sheet_navigation_signals(artifact_path, kwargs)
        if not signals:
            return BehavioralTestResult(
                test_name=self.test_name,
                dimension=self.dimension,
                format=self.format,
                passed=True,
                score=1.0,
                threshold=0.80,
                confidence=0.20,
                metadata={"applicable": False, "parser_support": "openpyxl", "sheet_names": []},
            )
        score = sum(signal.score for signal in signals) / len(signals)
        findings = [
            {
                "severity": "warning",
                "issue": issue,
                "sheet_name": signal.sheet_name,
                "sheet_index": signal.sheet_index,
                "content_terms": list(signal.content_terms),
            }
            for signal in signals
            for issue in _signal_issues(signal)
        ]
        answerer: BehavioralAnswerer | None = kwargs.get("answerer")
        questions = _navigation_questions(signals)
        answer_retention_metadata: dict[str, Any] = {
            "llm_answering_enabled": answerer is not None,
            "navigation_question_count": len(questions),
        }
        if answerer is not None and questions:
            context = _sheet_navigation_context(signals)
            retention = score_answer_retention(
                questions=questions,
                baseline_context=str(kwargs.get("baseline_text") or context),
                candidate_context=str(kwargs.get("candidate_text") or context),
                answerer=answerer,
            )
            score = min(score, retention.retention)
            answer_retention_metadata.update(
                {
                    "baseline_accuracy": retention.baseline_accuracy,
                    "candidate_accuracy": retention.candidate_accuracy,
                    "answer_accuracy_retention": retention.retention,
                }
            )
            findings.extend(retention.findings)
        return BehavioralTestResult(
            test_name=self.test_name,
            dimension=self.dimension,
            format=self.format,
            passed=score >= 0.80,
            score=round(score, 4),
            threshold=0.80,
            confidence=0.55,
            findings=findings,
            metadata={
                "applicable": True,
                "parser_support": "openpyxl",
                "sheet_names": [signal.sheet_name for signal in signals],
                "per_sheet": [
                    {
                        "sheet_name": signal.sheet_name,
                        "sheet_index": signal.sheet_index,
                        "state": signal.state,
                        "non_empty_cell_count": signal.non_empty_cell_count,
                        "content_terms": list(signal.content_terms),
                        "score": signal.score,
                        "issue": signal.issue,
                        "issues": list(_signal_issues(signal)),
                    }
                    for signal in signals
                ],
                **answer_retention_metadata,
            },
        )


def xlsx_sheet_navigation_signals(
    artifact_path: Path,
    kwargs: dict[str, Any] | None = None,
) -> list[XLSXSheetNavigationSignal]:
    """Return sheet-navigation signals from explicit names or workbook content."""
    kwargs = kwargs or {}
    if "sheet_names" in kwargs:
        sheet_names = _validate_sheet_names(kwargs["sheet_names"])
        return [
            _signal_from_name(name=name, sheet_index=index)
            for index, name in enumerate(sheet_names, start=1)
        ]
    if not artifact_path.exists():
        return []
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(str(artifact_path), read_only=True, data_only=True)
        try:
            signals: list[XLSXSheetNavigationSignal] = []
            for index, worksheet in enumerate(workbook.worksheets, start=1):
                content_terms, non_empty_cell_count = _sheet_content_terms(worksheet)
                if non_empty_cell_count == 0:
                    continue
                issues = _sheet_issues(
                    worksheet.title,
                    sheet_index=index,
                    state=getattr(worksheet, "sheet_state", "visible"),
                    content_terms=content_terms,
                )
                signals.append(
                    XLSXSheetNavigationSignal(
                        sheet_name=worksheet.title,
                        sheet_index=index,
                        state=getattr(worksheet, "sheet_state", "visible"),
                        non_empty_cell_count=non_empty_cell_count,
                        content_terms=content_terms,
                        issue=issues[0] if issues else "",
                        issues=issues,
                    )
                )
            return signals
        finally:
            workbook.close()
    except Exception:  # noqa: BLE001 - best-effort sheet navigation signal.
        return []


def _validate_sheet_names(value: Any) -> list[str]:
    if isinstance(value, (str, bytes)) or not isinstance(value, (list, tuple)):
        raise ValueError("sheet_names must be a list of non-empty strings")
    names: list[str] = []
    for name in value:
        if not isinstance(name, str) or not name.strip():
            raise ValueError("sheet_names must be a list of non-empty strings")
        names.append(name)
    return names


def _descriptive_sheet_name(name: str) -> bool:
    normalized = name.strip().lower()
    if not normalized:
        return False
    if normalized.startswith("sheet") and normalized[5:].isdigit():
        return False
    return len(normalized) >= 4


def _signal_from_name(name: str, sheet_index: int) -> XLSXSheetNavigationSignal:
    issues = _sheet_issues(
        name,
        sheet_index=sheet_index,
        state="visible",
        content_terms=(),
    )
    return XLSXSheetNavigationSignal(
        sheet_name=name,
        sheet_index=sheet_index,
        state="unknown",
        non_empty_cell_count=1,
        content_terms=(),
        issue=issues[0] if issues else "",
        issues=issues,
    )


def _sheet_issues(
    name: str,
    *,
    sheet_index: int,
    state: str,
    content_terms: tuple[str, ...],
) -> tuple[str, ...]:
    issues: list[str] = []
    descriptive = _descriptive_sheet_name(name)
    if not descriptive:
        issues.append("non_descriptive_sheet_name")
    if state != "visible":
        issues.append("data_sheet_hidden")
    if descriptive and content_terms and not (_tokens(name) & set(content_terms)):
        issues.append("sheet_name_purpose_unclear")
    if sheet_index > 1 and (_tokens(name) & _OVERVIEW_SHEET_TOKENS):
        issues.append("overview_sheet_not_first")
    return tuple(issues)


def _sheet_content_terms(worksheet: Any) -> tuple[tuple[str, ...], int]:
    counts: dict[str, int] = {}
    non_empty = 0
    for row in worksheet.iter_rows(values_only=True):
        for value in row:
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            non_empty += 1
            for token in _tokens(text):
                counts[token] = counts.get(token, 0) + 1
        if non_empty >= 60:
            break
    ranked = sorted(counts, key=lambda token: (-counts[token], token))
    return tuple(ranked[:8]), non_empty


def _navigation_questions(signals: list[XLSXSheetNavigationSignal]) -> list[GeneratedQuestion]:
    questions: list[GeneratedQuestion] = []
    for signal in signals:
        for term in signal.content_terms[:2]:
            questions.append(
                GeneratedQuestion(
                    question=f"Which worksheet contains information about {term}?",
                    expected_answer=signal.sheet_name,
                    source_dimension="sheet_organization",
                )
            )
            break
        if len(questions) >= 5:
            break
    return questions


def _sheet_navigation_context(signals: list[XLSXSheetNavigationSignal]) -> str:
    rows = []
    for signal in signals:
        purpose = ", ".join(signal.content_terms) or "no sampled content terms"
        rows.append(f"{signal.sheet_name}: {purpose}")
    return "\n".join(rows)


def _signal_issues(signal: XLSXSheetNavigationSignal) -> tuple[str, ...]:
    return signal.issues or ((signal.issue,) if signal.issue else ())


def _tokens(text: str) -> set[str]:
    tokens = set()
    for token in re.findall(r"[a-zA-Z][a-zA-Z0-9]{2,}", text.lower()):
        if token in _STOPWORDS:
            continue
        tokens.add(token)
    return tokens
