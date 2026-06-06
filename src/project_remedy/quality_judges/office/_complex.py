"""OOXML complex-content description helpers for Office quality judges."""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from defusedxml.ElementTree import fromstring as _safe_fromstring

from project_remedy._zip_safety import read_zip_member_safely
from project_remedy.behavioral_proxies.office.xlsx.alt_text_substitution import (
    _xlsx_drawing_objects,
)
from project_remedy.quality_judges.office._ooxml import (
    attr,
    is_docx_content_part,
    is_pptx_slide_part,
    local_name,
)
from project_remedy.quality_judges.shared._helpers import safe_ratio
from project_remedy.quality_judges.shared.base import QualityDimensionScore


COMPLEX_KEYWORDS = {
    "chart",
    "graph",
    "diagram",
    "equation",
    "formula",
    "plot",
    "model",
    "table",
}


@dataclass(frozen=True)
class ComplexOfficeObject:
    source: str
    object_index: int
    name: str
    title: str
    description: str
    kind: str = "drawing"

    @property
    def label(self) -> str:
        return " ".join(part for part in (self.name, self.title, self.description) if part).strip()

    @property
    def description_text(self) -> str:
        return " ".join(part for part in (self.title, self.description) if part).strip()


def docx_complex_objects(artifact_path: Path) -> list[ComplexOfficeObject]:
    objects = _complex_objects_from_ooxml(
        artifact_path,
        part_predicate=is_docx_content_part,
        target_local_names=("docPr", "cNvPr"),
    )
    objects.extend(
        _equation_objects(
            artifact_path,
            part_predicate=is_docx_content_part,
            start_index=len(objects) + 1,
        )
    )
    return objects


def pptx_complex_objects(artifact_path: Path) -> list[ComplexOfficeObject]:
    objects = _complex_objects_from_ooxml(
        artifact_path,
        part_predicate=is_pptx_slide_part,
        target_local_names=("cNvPr",),
    )
    objects.extend(
        _equation_objects(
            artifact_path,
            part_predicate=is_pptx_slide_part,
            start_index=len(objects) + 1,
        )
    )
    return objects


def xlsx_complex_objects(artifact_path: Path) -> list[ComplexOfficeObject]:
    objects = [
        ComplexOfficeObject(
            source=item.source,
            object_index=item.object_index,
            name=item.name,
            title=item.title,
            description=item.description,
            kind="drawing",
        )
        for item in _xlsx_drawing_objects(artifact_path)
    ]
    objects.extend(_xlsx_formula_objects(artifact_path, start_index=len(objects) + 1))
    return objects


def score_complex_objects(
    objects: list[ComplexOfficeObject],
    *,
    judge_id: str,
    judge_version: str,
    fmt: str,
) -> QualityDimensionScore:
    """Score candidate complex objects for data-level descriptions."""
    candidates = [item for item in objects if _looks_complex(item)]
    if not candidates:
        return QualityDimensionScore(
            dimension="complex_content",
            format=fmt,
            score=1.0,
            variance=0.0,
            per_criterion={"complex_object_parser_coverage": 0.0},
            judge_versions=[f"{judge_id}:{judge_version}"],
            sample_findings=[],
            confidence=0.30,
        )

    failures = [item for item in candidates if not _has_data_level_description(item)]
    score = (len(candidates) - len(failures)) / len(candidates)
    per_criterion = {"data_level_description": round(score, 4)}
    for kind, criterion in (("formula", "formula_context"), ("equation", "equation_context")):
        subset = [item for item in candidates if item.kind == kind]
        if not subset:
            continue
        subset_failures = sum(1 for item in subset if not _has_data_level_description(item))
        per_criterion[criterion] = safe_ratio(len(subset) - subset_failures, len(subset))
    findings = [
        {
            "severity": "warning",
            "issue": "thin_complex_content_description",
            "source": item.source,
            "object_index": item.object_index,
            "name": item.name,
            "title": item.title,
            "description": item.description,
            "kind": item.kind,
        }
        for item in failures
    ]
    return QualityDimensionScore(
        dimension="complex_content",
        format=fmt,
        score=round(score, 4),
        variance=0.0,
        per_criterion=per_criterion,
        judge_versions=[f"{judge_id}:{judge_version}"],
        sample_findings=findings[:5],
        confidence=0.55,
    )


def _complex_objects_from_ooxml(
    artifact_path: Path,
    *,
    part_predicate: Callable[[str], bool],
    target_local_names: tuple[str, ...],
) -> list[ComplexOfficeObject]:
    if not artifact_path.exists():
        return []

    objects: list[ComplexOfficeObject] = []
    try:
        with ZipFile(artifact_path) as package:
            for part_name in sorted(package.namelist()):
                if not part_predicate(part_name):
                    continue
                content = read_zip_member_safely(package, part_name)
                if content is None:
                    continue
                objects.extend(
                    _complex_objects_from_xml(
                        content,
                        source=part_name,
                        target_local_names=target_local_names,
                    )
                )
    except (BadZipFile, OSError):
        return []
    return objects


def _complex_objects_from_xml(
    content: bytes,
    *,
    source: str,
    target_local_names: tuple[str, ...],
) -> list[ComplexOfficeObject]:
    try:
        root = _safe_fromstring(content)
    except ElementTree.ParseError:
        return []

    objects: list[ComplexOfficeObject] = []
    targets = set(target_local_names)
    for element in root.iter():
        if local_name(element.tag) not in targets:
            continue
        objects.append(
            ComplexOfficeObject(
                source=source,
                object_index=len(objects) + 1,
                name=attr(element, "name"),
                title=attr(element, "title"),
                description=attr(element, "descr"),
            )
        )
    return objects


def _looks_complex(item: ComplexOfficeObject) -> bool:
    normalized = item.label.lower()
    return any(keyword in normalized for keyword in COMPLEX_KEYWORDS)


def _has_data_level_description(item: ComplexOfficeObject) -> bool:
    if item.kind in {"formula", "equation"}:
        return _formula_has_context(item.description)
    description = item.description_text
    if not description:
        return False
    return any(char.isdigit() for char in description) or len(description.split()) >= 8


def _equation_objects(
    artifact_path: Path,
    *,
    part_predicate: Callable[[str], bool],
    start_index: int,
) -> list[ComplexOfficeObject]:
    if not artifact_path.exists():
        return []
    objects: list[ComplexOfficeObject] = []
    try:
        with ZipFile(artifact_path) as package:
            for part_name in sorted(package.namelist()):
                if not part_predicate(part_name):
                    continue
                content = read_zip_member_safely(package, part_name)
                if content is None:
                    continue
                objects.extend(
                    _equation_objects_from_xml(
                        content,
                        source=part_name,
                        start_index=start_index + len(objects),
                    )
                )
    except (BadZipFile, OSError):
        return []
    return objects


def _equation_objects_from_xml(
    content: bytes,
    *,
    source: str,
    start_index: int,
) -> list[ComplexOfficeObject]:
    try:
        root = _safe_fromstring(content)
    except ElementTree.ParseError:
        return []

    objects: list[ComplexOfficeObject] = []
    for paragraph in root.iter():
        if local_name(paragraph.tag) != "p" or not _contains_math(paragraph):
            continue
        objects.append(
            ComplexOfficeObject(
                source=f"{source}#equation-{len(objects) + 1}",
                object_index=start_index + len(objects),
                name=f"Equation {len(objects) + 1}",
                title=_math_text(paragraph) or "equation",
                description=_paragraph_text_excluding_math(paragraph),
                kind="equation",
            )
        )
    return objects


def _contains_math(element: ElementTree.Element) -> bool:
    return any(
        local_name(child.tag) in {"oMath", "oMathPara"}
        for child in element.iter()
    )


def _math_text(element: ElementTree.Element) -> str:
    parts: list[str] = []
    for child in element.iter():
        if local_name(child.tag) not in {"oMath", "oMathPara"}:
            continue
        parts.extend(_text_values(child))
    return " ".join(part for part in parts if part)


def _paragraph_text_excluding_math(element: ElementTree.Element) -> str:
    parts = _text_values_excluding_math(element)
    return " ".join(part for part in parts if part)


def _text_values(element: ElementTree.Element) -> list[str]:
    values: list[str] = []
    for child in element.iter():
        if local_name(child.tag) == "t" and child.text:
            value = " ".join(child.text.split())
            if value:
                values.append(value)
    return values


def _text_values_excluding_math(element: ElementTree.Element) -> list[str]:
    if local_name(element.tag) in {"oMath", "oMathPara"}:
        return []
    if local_name(element.tag) == "t" and element.text:
        value = " ".join(element.text.split())
        return [value] if value else []
    values: list[str] = []
    for child in element:
        values.extend(_text_values_excluding_math(child))
    return values


def _xlsx_formula_objects(
    artifact_path: Path,
    *,
    start_index: int,
) -> list[ComplexOfficeObject]:
    if not artifact_path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        workbook = load_workbook(str(artifact_path), data_only=False)
    except Exception:  # noqa: BLE001 - malformed or unsupported workbook.
        return []
    try:
        objects: list[ComplexOfficeObject] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not _is_formula_cell(value):
                        continue
                    objects.append(
                        ComplexOfficeObject(
                            source=f"{worksheet.title}!{cell.coordinate}",
                            object_index=start_index + len(objects),
                            name=f"Formula {worksheet.title}!{cell.coordinate}",
                            title=str(value),
                            description=_formula_context(worksheet, cell),
                            kind="formula",
                        )
                    )
        return objects
    finally:
        workbook.close()


def _is_formula_cell(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _formula_context(worksheet: object, cell: object) -> str:
    labels: list[str] = []
    row = int(getattr(cell, "row", 0) or 0)
    column = int(getattr(cell, "column", 0) or 0)
    if column > 1:
        labels.append(_cell_text(worksheet.cell(row=row, column=column - 1)))
    if row > 1:
        labels.append(_cell_text(worksheet.cell(row=row - 1, column=column)))
    return " ".join(label for label in labels if label)


def _cell_text(cell: object) -> str:
    value = getattr(cell, "value", None)
    if _is_formula_cell(value):
        return ""
    return " ".join(str(value or "").split())


def _formula_has_context(description: str) -> bool:
    words = [
        word
        for word in description.split()
        if any(character.isalpha() for character in word)
    ]
    return bool(words)
